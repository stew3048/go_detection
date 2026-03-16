"""
將 30ep 與 100ep 兩個模型的推論結果匯出為 Excel
輸出：runs/predict_comparison.xlsx
"""

from pathlib import Path
from ultralytics import YOLO
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── 設定 ─────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).parent.resolve()
MODELS = {
    "30ep":  PROJECT_ROOT / "runs" / "go_black_white"       / "weights" / "best.pt",
    "100ep": PROJECT_ROOT / "runs" / "go_black_white_100ep" / "weights" / "best.pt",
}
TEST_DIR    = PROJECT_ROOT / "datasets" / "go-games-1" / "test" / "images"
OUTPUT_XLSX = PROJECT_ROOT / "runs" / "predict_comparison.xlsx"
CONF = 0.25
IOU  = 0.45
# ─────────────────────────────────────────────────────


def run_predict(model_path: Path) -> dict:
    """執行推論，回傳 {filename: {"black": n, "white": n, "conf_avg": f}} """
    model = YOLO(str(model_path))
    results = model.predict(
        source   = str(TEST_DIR),
        conf     = CONF,
        iou      = IOU,
        save     = False,
        verbose  = False,
    )
    data = {}
    for r in results:
        name = Path(r.path).name
        boxes = r.boxes
        if boxes is None or len(boxes) == 0:
            data[name] = {"black": 0, "white": 0, "total": 0, "conf_avg": 0.0}
            continue
        black = int((boxes.cls == 0).sum())
        white = int((boxes.cls == 1).sum())
        conf_avg = float(boxes.conf.mean()) if len(boxes.conf) > 0 else 0.0
        data[name] = {
            "black": black,
            "white": white,
            "total": black + white,
            "conf_avg": round(conf_avg, 3),
        }
    return data


def build_excel(data_30: dict, data_100: dict, output: Path):
    wb = openpyxl.Workbook()

    # ── 樣式定義 ────────────────────────────────────
    header_fill   = PatternFill("solid", fgColor="1F4E79")
    ep30_fill     = PatternFill("solid", fgColor="D6E4F0")
    ep100_fill    = PatternFill("solid", fgColor="D5F5E3")
    alt_fill      = PatternFill("solid", fgColor="F2F2F2")
    better_fill   = PatternFill("solid", fgColor="ABEBC6")   # 100ep 優於 30ep 時
    worse_fill    = PatternFill("solid", fgColor="FADBD8")   # 100ep 劣於 30ep 時
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    bold_font     = Font(bold=True)
    center        = Alignment(horizontal="center", vertical="center")
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_cell(cell, fill=None, font=None, align=center, border=thin_border):
        if fill:   cell.fill      = fill
        if font:   cell.font      = font
        cell.alignment = align
        cell.border    = border

    # ════════════════════════════════════════════════
    # 工作表 1：逐圖比較
    # ════════════════════════════════════════════════
    ws = wb.active
    ws.title = "逐圖比較"

    # 第一列：大標題
    titles = [
        ("A1:A2", "圖片名稱"),
        ("B1:D1", "30 Epoch 模型"),
        ("E1:G1", "100 Epoch 模型"),
        ("H1:J1", "差異（100ep - 30ep）"),
    ]
    for cell_range, text in titles:
        ws.merge_cells(cell_range)
        c = ws[cell_range.split(":")[0]]
        c.value = text
        style_cell(c, fill=header_fill, font=header_font)

    # 第二列：子標題（跳過 col=1，因為 A1:A2 已合併）
    sub = {2: "黑棋", 3: "白棋", 4: "合計", 5: "黑棋", 6: "白棋", 7: "合計", 8: "△黑棋", 9: "△白棋", 10: "△合計"}
    for col, val in sub.items():
        c = ws.cell(row=2, column=col, value=val)
        style_cell(c, fill=PatternFill("solid", fgColor="2E75B6"), font=Font(bold=True, color="FFFFFF", size=10))

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # 資料列
    filenames = sorted(set(data_30) | set(data_100))
    for i, fname in enumerate(filenames):
        row = i + 3
        d30  = data_30.get(fname,  {"black": 0, "white": 0, "total": 0})
        d100 = data_100.get(fname, {"black": 0, "white": 0, "total": 0})
        delta_b = d100["black"] - d30["black"]
        delta_w = d100["white"] - d30["white"]
        delta_t = d100["total"] - d30["total"]

        row_fill = alt_fill if i % 2 == 0 else None

        # 圖片名稱（縮短顯示）
        short_name = fname[:45] + "..." if len(fname) > 45 else fname
        c = ws.cell(row=row, column=1, value=short_name)
        style_cell(c, fill=row_fill, align=Alignment(horizontal="left", vertical="center"), border=thin_border)

        # 30ep 數據
        for col, val in zip([2, 3, 4], [d30["black"], d30["white"], d30["total"]]):
            c = ws.cell(row=row, column=col, value=val)
            style_cell(c, fill=ep30_fill)

        # 100ep 數據
        for col, val in zip([5, 6, 7], [d100["black"], d100["white"], d100["total"]]):
            c = ws.cell(row=row, column=col, value=val)
            style_cell(c, fill=ep100_fill)

        # 差異欄（有變化時標色）
        for col, delta in zip([8, 9, 10], [delta_b, delta_w, delta_t]):
            c = ws.cell(row=row, column=col, value=delta)
            diff_fill = better_fill if delta > 0 else (worse_fill if delta < 0 else row_fill)
            style_cell(c, fill=diff_fill)

    # 合計列
    total_row = len(filenames) + 3
    ws.cell(total_row, 1, "合計").font = bold_font
    sums = {
        "30_b": sum(d["black"] for d in data_30.values()),
        "30_w": sum(d["white"] for d in data_30.values()),
        "30_t": sum(d["total"] for d in data_30.values()),
        "100_b": sum(d["black"] for d in data_100.values()),
        "100_w": sum(d["white"] for d in data_100.values()),
        "100_t": sum(d["total"] for d in data_100.values()),
    }
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    vals = [
        sums["30_b"], sums["30_w"], sums["30_t"],
        sums["100_b"], sums["100_w"], sums["100_t"],
        sums["100_b"] - sums["30_b"],
        sums["100_w"] - sums["30_w"],
        sums["100_t"] - sums["30_t"],
    ]
    for col, val in enumerate(vals, 2):
        c = ws.cell(total_row, col, val)
        style_cell(c, fill=total_fill, font=bold_font)

    # 欄寬
    ws.column_dimensions["A"].width = 48
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 10
    ws.freeze_panes = "B3"

    # ════════════════════════════════════════════════
    # 工作表 2：模型比較摘要
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet("模型比較摘要")
    summary = [
        ["指標",        "30 Epoch", "100 Epoch", "進步幅度"],
        ["mAP50",       0.719,       0.940,       None],
        ["mAP50-95",    0.599,       0.816,       None],
        ["Precision",   0.939,       0.963,       None],
        ["Recall",      0.707,       0.925,       None],
        ["訓練時間",    "4.73 hrs",  "15.98 hrs", "-"],
        ["偵測黑棋總數", sums["30_b"], sums["100_b"], None],
        ["偵測白棋總數", sums["30_w"], sums["100_w"], None],
        ["偵測棋子總數", sums["30_t"], sums["100_t"], None],
    ]
    # 計算進步幅度
    for row_data in summary[1:]:
        if row_data[3] is None and isinstance(row_data[1], float):
            row_data[3] = f"+{round((row_data[2] - row_data[1]) / row_data[1] * 100, 1)}%"
        elif row_data[3] is None and isinstance(row_data[1], int):
            diff = row_data[2] - row_data[1]
            row_data[3] = f"+{diff}" if diff >= 0 else str(diff)

    for r_idx, row_data in enumerate(summary, 1):
        for c_idx, val in enumerate(row_data, 1):
            c = ws2.cell(r_idx, c_idx, val)
            if r_idx == 1:
                style_cell(c, fill=header_fill, font=header_font)
            else:
                fill = ep30_fill if c_idx == 2 else (ep100_fill if c_idx == 3 else None)
                style_cell(c, fill=fill, font=bold_font if c_idx == 1 else None)

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 12
    ws2.row_dimensions[1].height = 20

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Excel saved: {output}")


if __name__ == "__main__":
    print("[1/3] Running 30ep model inference...")
    data_30 = run_predict(MODELS["30ep"])

    print("[2/3] Running 100ep model inference...")
    data_100 = run_predict(MODELS["100ep"])

    print("[3/3] Building Excel...")
    build_excel(data_30, data_100, OUTPUT_XLSX)
    print("Done!")
